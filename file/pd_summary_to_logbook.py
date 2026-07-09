import sys
import argparse
import functools
from pathlib import Path

print = functools.partial(print, flush=True)

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
import openpyxl
from tools import get_paths, get_excluded_trackids

STAT_COLS = ['n', 'min', 'q25', 'median', 'q75', 'max', 'iqr', 'mean', 'std', 'se', 'skewness']


def reconstruct_trackid(substrate_no, track_no, layer):
    """Reconstruct trackid from raw columns, matching the Excel CONCATENATE formula."""
    if substrate_no is None or track_no is None:
        return None
    base = f'0{int(substrate_no):03d}_0{int(track_no)}'
    if layer is not None and int(layer) != 1:
        base += f'_0{int(layer)}'
    return base


def build_trackid_row_map(ws, substrate_col, track_col, layer_col):
    """Return dict of {trackid: [row_numbers]} reconstructed from raw source columns."""
    mapping = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        substrate_no = row[substrate_col - 1]
        track_no = row[track_col - 1]
        layer = row[layer_col - 1]
        trackid = reconstruct_trackid(substrate_no, track_no, layer)
        if trackid is None:
            continue
        mapping.setdefault(trackid, []).append(row_num)
    return mapping


STAT_UNITS = {
    'n':        '',
    'min':      ' [bits]',
    'q25':      ' [bits]',
    'median':   ' [bits]',
    'q75':      ' [bits]',
    'max':      ' [bits]',
    'iqr':      ' [bits]',
    'mean':     ' [bits]',
    'std':      ' [bits]',
    'se':       ' [bits]',
    'skewness': '',
}


def get_or_create_pd_columns(ws, pd_n):
    """Return {stat: col_index} for PD_N headers; create headers if missing."""
    target_headers = {stat: f'PD_{pd_n}_{stat}{STAT_UNITS[stat]}' for stat in STAT_COLS}
    header_row = next(ws.iter_rows(min_row=1, max_row=1))

    col_map = {}
    for cell in header_row:
        if cell.value in target_headers.values():
            stat = next(s for s, h in target_headers.items() if h == cell.value)
            col_map[stat] = cell.column

    missing = [s for s in STAT_COLS if s not in col_map]
    if missing:
        next_col = ws.max_column + 1
        for stat in missing:
            ws.cell(row=1, column=next_col, value=target_headers[stat])
            col_map[stat] = next_col
            next_col += 1
        print(f'  Created {len(missing)} new header(s): {[target_headers[s] for s in missing]}')

    return col_map


def find_source_columns(ws):
    """Return (substrate_col, track_col, layer_col) 1-based indices."""
    targets = {'Substrate No.': None, 'Track No.': None, 'Layer': None}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value in targets:
            targets[cell.value] = cell.column
    missing = [k for k, v in targets.items() if v is None]
    if missing:
        raise ValueError(f'Logbook header columns not found: {missing}')
    return targets['Substrate No.'], targets['Track No.'], targets['Layer']


def main():
    parser = argparse.ArgumentParser(
        description='Write PD summary CSV statistics into the Excel logbook.')
    parser.add_argument('--pd', type=int, choices=[1, 2], default=2,
                        help='Photodiode number to process (default: 2)')
    parser.add_argument('--dry_run', action='store_true',
                        help='Print what would be written without saving')
    parser.add_argument('--output', type=str, default=None,
                        help='Save to a different path instead of overwriting the logbook')
    args = parser.parse_args()

    paths = get_paths()
    hdf5_dir = paths['hdf5']
    logbook_path = paths['logbook']

    csv_name = f'AMPM_Photodiode{args.pd}Bits_summary.csv'
    csv_path = hdf5_dir / csv_name

    if not csv_path.exists():
        print(f'Error: CSV not found at {csv_path}')
        print(f'Run meas/summarise_timeseries.py with group=AMPM, '
              f'series=Photodiode{args.pd}Bits first.')
        sys.exit(1)

    print(f'Reading {csv_name}...')
    df = pd.read_csv(csv_path)
    df['trackid'] = df['trackid'].astype(str).str.strip()

    excluded = set(get_excluded_trackids())
    n_before = len(df)
    df = df[~df['trackid'].isin(excluded)]
    n_excluded = n_before - len(df)
    if n_excluded:
        print(f'  Excluded {n_excluded} trackid(s) with corrupted PD data')

    csv_index = df.set_index('trackid')
    print(f'  {len(csv_index)} trackids available to write')

    print(f'Loading logbook: {logbook_path.name}...')

    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb_write = openpyxl.load_workbook(logbook_path)
    ws_write = wb_write['Logbook']

    # Build trackid->row map from raw source columns (avoids formula cache issues)
    substrate_col, track_col, layer_col = find_source_columns(ws_write)
    trackid_row_map = build_trackid_row_map(ws_write, substrate_col, track_col, layer_col)

    print(f'  {len(trackid_row_map)} unique trackids found in logbook')
    col_map = get_or_create_pd_columns(ws_write, args.pd)
    print(f'  Column positions: { {stat: ws_write.cell(1, c).value for stat, c in col_map.items()} }')

    n_written = 0
    n_skipped_no_match = 0
    n_skipped_excluded = 0

    for trackid, csv_row in csv_index.iterrows():
        if trackid in excluded:
            n_skipped_excluded += 1
            continue

        if trackid not in trackid_row_map:
            n_skipped_no_match += 1
            continue

        for excel_row in trackid_row_map[trackid]:
            for stat in STAT_COLS:
                value = round(float(csv_row[stat]), 3)
                col = col_map[stat]
                if args.dry_run:
                    header = ws_write.cell(1, col).value
                    print(f'  [dry_run] row {excel_row} trackid={trackid}: {header} = {value}')
                else:
                    ws_write.cell(row=excel_row, column=col, value=value)

        n_written += 1

    print(f'\nSummary:')
    print(f'  Rows updated:             {n_written}')
    print(f'  Skipped (no CSV match):   {n_skipped_no_match}')
    print(f'  Skipped (excluded):       {n_skipped_excluded}')

    if args.dry_run:
        print('\n[dry_run] No changes saved.')
    else:
        out_path = Path(args.output) if args.output else logbook_path
        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                wb_write.save(out_path)
            print(f'\nSaved: {out_path}')
        except PermissionError:
            print(f'\nError: Could not save to {out_path}')
            print('The logbook may be open in Excel — close it and try again.')
            print('Or use --output path/to/output.xlsx to save a copy.')
            sys.exit(1)


if __name__ == '__main__':
    main()
